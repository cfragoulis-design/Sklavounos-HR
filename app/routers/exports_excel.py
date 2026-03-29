from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import extract, select
from sqlalchemy.orm import Session, joinedload

from app.dependencies import get_current_admin, get_db
from app.models import LeaveRequest

router = APIRouter(prefix="/exports", tags=["exports"])


@router.get("/payroll-excel")
def export_payroll_excel(
    year: int = Query(...),
    month: int = Query(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    stmt = (
        select(LeaveRequest)
        .options(joinedload(LeaveRequest.employee), joinedload(LeaveRequest.leave_type))
        .where(
            LeaveRequest.status == "approved",
            extract("year", LeaveRequest.date_from) == year,
            extract("month", LeaveRequest.date_from) == month,
        )
        .order_by(LeaveRequest.date_from.asc(), LeaveRequest.id.asc())
    )

    rows = db.execute(stmt).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Leave"
    ws.freeze_panes = "A2"

    headers = ["Employee", "Leave Type", "Date From", "Date To", "Days"]
    ws.append(headers)

    center_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    thin_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for column_index in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=column_index)
        cell.font = Font(bold=True)
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    for row_index, record in enumerate(rows, start=2):
        ws.cell(row=row_index, column=1, value=record.employee.full_name if record.employee else "")
        ws.cell(row=row_index, column=2, value=record.leave_type.name if record.leave_type else "")

        date_from_cell = ws.cell(row=row_index, column=3, value=record.date_from)
        date_to_cell = ws.cell(row=row_index, column=4, value=record.date_to)
        days_cell = ws.cell(
            row=row_index,
            column=5,
            value=f'=IF(AND(C{row_index}<>"",D{row_index}<>""),D{row_index}-C{row_index}+1,"")',
        )

        date_from_cell.number_format = "dd/mm/yyyy"
        date_to_cell.number_format = "dd/mm/yyyy"
        days_cell.number_format = "0"

        for column_index in range(1, len(headers) + 1):
            cell = ws.cell(row=row_index, column=column_index)
            cell.alignment = center_alignment
            cell.border = thin_border

    if rows:
        last_data_row = len(rows) + 1
        table = Table(displayName="PayrollLeaveTable", ref=f"A1:E{last_data_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
        ws.auto_filter.ref = f"A1:E{last_data_row}"

    desired_widths = {
        "A": 24,
        "B": 18,
        "C": 14,
        "D": 14,
        "E": 10,
    }
    for column_letter, width in desired_widths.items():
        ws.column_dimensions[column_letter].width = width

    for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 2), min_col=1, max_col=5):
        for cell in row:
            cell.alignment = center_alignment

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"payroll_leave_{year}_{month}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
